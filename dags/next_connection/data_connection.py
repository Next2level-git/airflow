import time
import io
import pandas as pd
import psycopg2
import sqlalchemy
from openpyxl import load_workbook
from io import BytesIO
from pyxlsb import open_workbook
from abc import ABC
from concurrent.futures import ThreadPoolExecutor
from functools import partial
from .logger import INFO, get_logger
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.runtime.client_request_exception import ClientRequestException
from office365.runtime.client_request import HTTPError
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File


logger = get_logger(__name__, INFO)


class DataConnection(ABC):

    def create_connection(self):
        """Function to create a data connection"""
        pass

    def execute_query(self, query):
        """Function to execute a query"""
        pass

    def update_query(self, query, values):
        """Function to execute a update query with multiple values"""
        pass

    def execute_insert(self, data_frame, table_name, truncate):
        """Function to insert data.
        'action' will indicate if we must replace, append info to the table. Options are fail, replace or append
        """
        pass

    def execute_upsert(self, data_frame, table_name, primary_keys):
        """Function to insert/update (upsert) data.
        'action' will indicate if we must replace, append info to the table. Options are fail, replace or append
        """
        pass

    def destroy_connection(self):
        """Function to destroy an active connection"""
        pass

    def google_sheet_to_df(self, sheet_name):
        """Function to load an dataframe from google sheet"""
        pass


class OneDriveConnection:

    def __init__(self, username, password):
        self.username = username
        self.password = password
        self.object_list = []
        self.path_sep = "/"

    def move_to(self, url, old_url, new_url):
        """
        Move one file to new url
        :type url: str
        :type old_url: str
        :type new_url: str
        """
        ctx_auth = AuthenticationContext(url)
        ctx_auth.acquire_token_for_user(self.username, self.password)
        ctx = ClientContext(url, ctx_auth)
        file = ctx.web.get_file_by_server_relative_url(old_url)
        file.moveto(new_relative_url=new_url, flag=1)
        ctx.execute_query()

    def get_files_list_onedrive(
        self, url, folderpath="", item_childrem_type=["Files", "Folders"]
    ):
        ctx_auth = AuthenticationContext(url)
        ctx_auth.acquire_token_for_user(self.username, self.password)
        ctx = ClientContext(url, ctx_auth)
        root_folder = ctx.web.get_folder_by_server_relative_url(folderpath)

        result_list = self.enum_folder(root_folder, item_childrem_type, self.print_file)

        # items = doc_lib.items.select(["FileSystemObjectType"]).expand(["File", "Folder"]).get().execute_query()
        return result_list
        # else:
        # logger.error(f"Error, File Not Found. Status code: {response.status_code}, Reason: {response.reason}")
        # raise response

    def enum_folder(self, parent_folder, item_childrem_type, fn):
        """
        :type parent_folder: Folder
        :type item_childrem_type: str
        :type fn: (File)-> None
        """
        parent_folder.expand(item_childrem_type).get().execute_query()
        for file in parent_folder.files:  # type: File
            fn(file)

        for folder in parent_folder.folders:  # type: Folder
            fn(folder)
            self.enum_folder(folder, item_childrem_type, fn)

    def print_file(self, f):
        self.object_list.append(f.properties["ServerRelativeUrl"])
        print(f.properties["ServerRelativeUrl"])

    def excel_drive_to_df(self, url, filepath, sheetname=None, skiprows=0, engine=None):
        ctx_auth = AuthenticationContext(url)
        ctx_auth.acquire_token_for_user(self.username, self.password)
        ctx = ClientContext(url, ctx_auth)
        response = File.open_binary(ctx, filepath)
        if response.status_code == 200:
            bytes_io = BytesIO(response.content)
            return pd.read_excel(
                bytes_io, sheet_name=sheetname, skiprows=skiprows, engine=engine
            )
        else:
            logger.error(
                f"Error, File Not Found. Status code: {response.status_code}, Reason: {response.reason}"
            )
            raise Exception(
                f"Error, File Not Found. Status code: {response.status_code}, Reason: {response.reason}"
            )

    def excel_upload_file_df(
        self, url, data, filepath, filename, engine, encoding="utf-8"
    ):
        try:
            ctx_auth = AuthenticationContext(url)
            ctx_auth.acquire_token_for_user(self.username, self.password)
            ctx = ClientContext(url, ctx_auth)

            buffer = BytesIO()
            writer_excel = pd.ExcelWriter(buffer, engine=engine)
            data.to_excel(
                writer_excel, sheet_name="info", index=False, encoding=encoding
            )
            writer_excel.save()

            target_folder = ctx.web.get_folder_by_server_relative_url(filepath)
            target_folder.upload_file(filename + ".xlsx", buffer.getvalue())
            ctx.execute_query()
            logger.info("Uploaded file with name: {}".format(filename + ".xlsx"))
        except Exception as e:
            logger.error("An error has occurred during process. ", e)
            raise

    def excel_upload_file_df_multiple_sheets(
        self, url, data, filepath, filename, engine, encoding="utf-8"
    ):
        """
        Method to creates an excel file with multiple sheets and upload it to OneDrive

        Args:
            url (str): URL from OneDrive
            data (dict): Data to create the excel file. Each key (str) is the sheetname and each value (Dataframe)
                         is the data to upload
            filepath (str): Path in OneDrive where the file is going to be created.
            filename (str): Name of the file, with extension.
            engine (str): Excel engine to use. Recommended 'xlsxwriter'
            encoding (str): Encoding of file. Default 'utf-8'
        Raises:
            RuntimeError: Error creating the file
        Returns:
            NA
        """
        try:
            ctx_auth = AuthenticationContext(url)
            ctx_auth.acquire_token_for_user(self.username, self.password)
            ctx = ClientContext(url, ctx_auth)

            buffer = BytesIO()
            writer_excel = pd.ExcelWriter(buffer, engine=engine)
            for key, value in data.items():
                value.to_excel(
                    writer_excel, sheet_name=key, index=False, encoding=encoding
                )
            writer_excel.save()

            target_folder = ctx.web.get_folder_by_server_relative_url(filepath)
            target_folder.upload_file(filename, buffer.getvalue())
            ctx.execute_query()
            logger.info("Uploaded file with name: {}".format(filename))
        except Exception as e:
            logger.error("An error has occurred during process. ", e)
            raise

    def create_folder_if_not_exists(
        self, url: str, base_path: str, folder_name: str, max_retry=2
    ):
        """
        Method to create OneDrive folder if not exists
        Args:
            url (str): URL from OneDrive
            base_path (str): path in the URL where folder should be created
            folder_name (str): name of the folder
            client_ctx (ClientContext): from office365
            max_retry: number of times to retry the request
        Returns:
            NA
        """
        full_path = base_path + self.path_sep + folder_name

        try:
            ctx_auth = AuthenticationContext(url)
            ctx_auth.acquire_token_for_user(self.username, self.password)
            ctx = ClientContext(url, ctx_auth)

            target_folder = (
                ctx.web.get_folder_by_server_relative_url(full_path)
                .get()
                .execute_query()
            )
        except (ClientRequestException, HTTPError) as e:
            if e.response.status_code == 404:
                base_folder = ctx.web.get_folder_by_server_relative_url(base_path)
                base_folder.folders.add(folder_name)
                ctx.execute_query_retry(max_retry=max_retry)

    def upload_file(
        self,
        url: str,
        base_path: str,
        file_data: list = list(),
        max_retry_new_file: int = 2,
        wait_time_sleep: float = 0.0,
        max_retry_new_folder: int = 2,
    ):
        """
        Method to create OneDrive folder if not exists
        Args:
            url (str): URL from OneDrive
            base_path (str): path in the URL where file should be uploaded
            client_ctx (ClientContext): from office365
            file_data (list): [file_contents (BytesIO), folder_name (str), file_name (str)]
            max_retry_new_file: number of times to retry the request of file upload
            wait_time_sleep: seconds (or fraction of) to sleep after file creation (in case of high concurrency)
            max_retry_new_folder: number of times to retry the request of folder creation
        Returns:
            NA
        """
        file_name = file_data[-1]
        folder_name = file_data[-2]
        file_contents = file_data[0]
        full_path = base_path

        try:
            ctx_auth = AuthenticationContext(url)
            ctx_auth.acquire_token_for_user(self.username, self.password)
            ctx = ClientContext(url, ctx_auth)
        except Exception as ctx_exc:
            logger.error(
                f"||* {type(ctx_exc).__name__} *|| getting client context for uploading file"
            )
            raise

        if len(folder_name) > 0:
            self.create_folder_if_not_exists(
                url, base_path, folder_name, max_retry_new_folder
            )
            full_path = full_path + self.path_sep + folder_name

        try:
            target_folder = ctx.web.get_folder_by_server_relative_url(full_path)
            target_folder.upload_file(file_name, file_contents)
            ctx.execute_query_retry(max_retry=max_retry_new_file)
            time.sleep(wait_time_sleep)
        except Exception as e:
            logger.error(
                f"/X\\(°-°)/X\\ ||* {type(e).__name__} *|| uploading {file_name} /X\\(°-°)/X\\"
            )
            raise

    def upload_file_list(
        self,
        url: str,
        base_path: str,
        files: list,
        num_workers: int = 1,
        max_retry_new_file: int = 2,
        wait_secs_post_new_file: float = 0.0,
        max_retry_new_folder: int = 2,
        pool_timeout: int = 300,
    ):
        """
        Method to upload a list of files
        Args:
            url (str): URL from OneDrive
            base_path (str): path in the URL where file should be uploaded
            files (list of lists): [[file_contents (BytesIO), folder_name (str), file_name (str)], ...]
            num_workers: number of worker processes to which jobs can be submitted (for concurrent execution)
            max_retry_new_file: number of times to retry the request of file upload
            wait_secs_post_new_file: seconds (or fraction of) to sleep after file creation (in case of high concurrency)
            max_retry_new_folder: number of times to retry the request of folder creation
            pool_timeout: timeout parameter (in seconds) for ThreadPoolExecutor
        Returns:
            NA
        """
        num_workers = num_workers

        if len(files) > 0:
            upload_bin_file = partial(
                self.upload_file,
                url,
                base_path,
                max_retry_new_file=max_retry_new_file,
                wait_time_sleep=wait_secs_post_new_file,
                max_retry_new_folder=max_retry_new_folder,
            )

            with ThreadPoolExecutor(max_workers=num_workers) as tpool_exec:
                try:
                    results = tpool_exec.map(
                        upload_bin_file, files, timeout=pool_timeout
                    )
                    lst_results = list(results)
                except Exception as e:
                    logger.error(
                        f"||* {type(e).__name__} *|| processing file list to OneDrive/Sharepoint"
                    )
                    raise

            logger.info(f"{len(files)} files uploaded to OneDrive/Sharepoint")

    def excelxlsb_drive_to_df(
        self,
        url,
        filepath,
        numbersheet=None,
        start_cell=None,
        end_cell=None,
        engine=None,
        headers=None,
    ):
        ctx_auth = AuthenticationContext(url)
        ctx_auth.acquire_token_for_user(self.username, self.password)
        ctx = ClientContext(url, ctx_auth)
        response = File.open_binary(ctx, filepath)

        if response.status_code == 200:
            if filepath.endswith(".xlsb"):
                # Si es un archivo XLSB
                with open_workbook(BytesIO(response.content)) as wb:
                    with wb.get_sheet(numbersheet) as sheet:
                        data = []
                        for row in sheet.rows():
                            row_data = [item.v for item in row]
                            data.append(row_data)

                headers = data[headers]
                df = pd.DataFrame(data[1:], columns=headers)
                df.dropna(axis=1, how="all", inplace=True)
                df.dropna(axis=0, how="all", inplace=True)
            else:
                # Si es un archivo XLSX (usando openpyxl)
                excel_data = response.content
                excel_workbook = load_workbook(
                    filename=BytesIO(excel_data), read_only=True, data_only=True
                )
                if sheetname is None:
                    sheet = excel_workbook.active
                else:
                    sheet = excel_workbook[sheetname]
                data = []
                for row in sheet.iter_rows(
                    min_row=start_cell[0],
                    max_row=end_cell[0],
                    min_col=start_cell[1],
                    max_col=end_cell[1],
                ):
                    row_data = [cell.value for cell in row]
                    data.append(row_data)

                headers = data[3]  # Assuming the first row contains headers
                df = pd.DataFrame(data[1:], columns=headers)
                df.dropna(axis=1, how="all", inplace=True)
                df.dropna(axis=0, how="all", inplace=True)

            return df
        else:
            logger.error(
                f"Error, File Not Found. Status code: {response.status_code}, Reason: {response.reason}"
            )
            raise Exception(
                f"Error, File Not Found. Status code: {response.status_code}, Reason: {response.reason}"
            )


class PostgreSQLConnection(DataConnection):
    """Creates and manages a PostgreSQL connection"""

    def __init__(
        self,
        user,
        password,
        host_name,
        database,
        schema="public",
        engine=None,
        connection=None,
    ):
        self.user = user
        self.password = password
        self.host_name = host_name
        self.database = database
        self.schema = schema
        self.engine = engine
        self.connection = connection

    def create_connection(self):
        try:
            self.engine = sqlalchemy.create_engine(
                "postgresql+psycopg2://{}:{}@{}/{}".format(
                    self.user, self.password, self.host_name, self.database
                )
            )
            self.connection = self.engine.connect()
        except Exception as ex:
            print("Exception while creating the postgresql connection {}".format(ex))
            raise

    def update_query(self, query, values):
        try:
            conn = self.engine.connect()
            cursor = conn.connection.cursor()

            psycopg2.extras.execute_values(
                cursor, query.as_string(cursor), values, template=None, page_size=100
            )
            conn.connection.commit()

        except Exception as ex:
            print("Exception while executing the update query {}").format(ex)
            raise

    def execute_query(self, query):
        try:
            return self.connection.execute(query)
        except Exception as ex:
            print("Exception while executing the postgresql query {}".format(ex))
            raise

    def execute_upsert(self, data_frame, schema, table_name, primary_keys, sep="|"):
        """
        Function to execute an upsert in a postgresql database given a pandas dataframe
        @type data_frame: pandas.DataFrame
        @type table_name: str
        @type primary_keys: list(tuple(str)) ej: [('pk_one', 'integer'), ('pk_two', 'varchar(16)')]
        """
        to_be_deleted = data_frame[[x[0] for x in primary_keys]].drop_duplicates()
        if len(to_be_deleted.index) < len(data_frame.index):
            raise ValueError(
                "Primary key constraint is violated in passed `data_frame`."
            )
        del_stream = io.StringIO()
        del_stream.write(
            to_be_deleted.to_csv(sep=sep, encoding="utf8", index=False, header=False)
        )
        del_stream.seek(0)
        try:
            conn = self.engine.connect()
            curs = conn.connection.cursor()
            curs.execute(
                """CREATE TEMP TABLE to_be_deleted_{0}
                            (
                                {1}
                            )""".format(
                    table_name, ",\n".join(["{} {}".format(*x) for x in primary_keys])
                )
            )
            curs.copy_from(del_stream, "to_be_deleted_{0}".format(table_name), sep=sep)
            curs.execute(
                """DELETE FROM
                                {0}.{1}
                            WHERE
                            ({2}) IN (SELECT {2} FROM to_be_deleted_{1})
                            """.format(
                    schema, table_name, ", ".join([x[0] for x in primary_keys])
                )
            )
            self.execute_insert(
                data_frame=data_frame,
                schema=schema,
                table_name=table_name,
                truncate=False,
                sep=sep,
            )
            conn.connection.commit()
        except Exception as exception:
            print(
                "Failure while upsert data to the table {}. {}".format(
                    table_name, exception
                )
            )
            raise exception
        finally:
            if curs is not None:
                curs.close()
            if conn is not None:
                conn.connection.close()

    def execute_insert(self, data_frame, schema, table_name, truncate=False, sep="|"):
        """Function to execute an insert in a postgresql database given a pandas dataframe"""
        try:

            pandas_sql_engine = pd.io.sql.pandasSQL_builder(self.engine, schema=schema)
            table = pd.io.sql.SQLTable(
                name=table_name,
                pandas_sql_engine=pandas_sql_engine,
                frame=data_frame,
                index=False,
                schema=schema,
                if_exists="fail",
            )

            columns = (
                str(list(data_frame.columns))
                .replace("[", "")
                .replace("]", "")
                .replace("'", "")
            )

            if not table.exists():
                table.create()
            if truncate:
                self.execute_query("TRUNCATE TABLE {}.{}".format(schema, table_name))
            string_data_io = io.StringIO()
            data_frame.to_csv(string_data_io, sep=sep, index=False)
            string_data_io.seek(0)
            with self.engine.connect() as connection:
                try:
                    with connection.connection.cursor() as cursor:
                        copy_cmd = (
                            "COPY %s.%s (%s) FROM STDIN HEADER DELIMITER '%s' CSV"
                            % (
                                schema,
                                table_name,
                                columns,
                                sep,
                            )
                        )
                        cursor.copy_expert(copy_cmd, string_data_io)
                    connection.connection.commit()
                finally:
                    if connection is not None:
                        connection.connection.close()
                    if cursor is not None:
                        cursor.close()

        except ValueError as exception:
            print(
                "Failure while inserting data to the table {}. {}".format(
                    table_name, exception
                )
            )
            raise

    def create_schema(self, schema: str) -> bool:
        """Function with the single responsability of create the schema in
        a particular Postgrsql connection

        Args:
            schema ([str]): Schema name

        Returns:
            [str]: True if created or existent
        """

        create_schema = f"""
            CREATE SCHEMA {schema};
        """

        has_schema = self.has_schema(schema)
        if not has_schema:
            _ = self.connection.execute(create_schema)
        return has_schema

    def has_schema(self, schema: str) -> bool:
        """Checks if the schema exists

        Args:
            schema (str): Schema name

        Returns:
            [bool]: True if schema is exists
        """
        has_schema = self.connection.engine.dialect.has_schema(
            schema=schema, connection=self.connection
        )

        return has_schema

    def create_table(
        self,
        schema: str,
        table_name: str,
        df_signature: pd.DataFrame,
        exists: bool = False,
    ) -> bool:
        """Create a table in the particular schema

        Args:
            schema (str): Schema name
            table_name (str): Table Name
            df_signature (DataFrame): Dataframe signature
            exists (bool, optional): boolean if the table exists (has_table). Defaults to False.

        Returns:
            [bool]: True if table is created
        """
        exists = self.has_table(schema, table_name)
        if not exists:
            df_signature.to_sql(
                table_name,
                self.connection,
                schema=schema.lower(),
                if_exists="append",
                index=False,
            )
            exists = True
        else:
            exists = False

        return exists

    def has_table(self, schema: str, table_name: str) -> bool:
        """Answer if the table exists

        Args:
            schema (str): Schema name
            table_name (str): Table Name

        Returns:
            [bool]: True if table is exists
        """
        exists = self.connection.engine.has_table(table_name, schema=schema)

        return exists

    def destroy_connection(self):
        try:
            if self.connection is not None:
                self.connection.close()
        except Exception as ex:
            print("Exception while destroying the postgresql connection {}".format(ex))
            raise
